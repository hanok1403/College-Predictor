from openpyxl import load_workbook
import numpy as np



def caste_to_col(caste,gender):
     column=1
     if(caste==1):
          column=10
          if(gender=="female"):
               column=11
     elif(caste==2):
          column=12
          if(gender=="female"):
               column=13
     elif(caste==3):
          column=14
          if(gender=="female"):
               column=15
     elif(caste==4):
          column=16
          if(gender=="female"):
               column=17
     elif(caste==5):
          column=18
          if(gender=="female"):
               column=19
     elif(caste==6):
          column=20
          if(gender=="female"):
               column=21
     elif(caste==7):
          column=22
          if(gender=="female"):
               column=23
     elif(caste==8):
          column=24
          if(gender=="female"):
               column=25
     else:
          column=26
          if(gender=="female"):
               column=27
     return column

def rankCalculator(rank,gender,course,caste):
     wb_2021_1 = load_workbook(r"TSEAMCET_2021.xlsx")
     wb_2020_3 = load_workbook(r"TSEAMCET_2020.xlsx")
     column=caste_to_col(caste,gender)
     sheet_1 = wb_2021_1[course]
     sheet_2 = wb_2020_3[course]
     row_count_1 = sheet_1.max_row
     row_count_2 = sheet_2.max_row
     dict1=dict()
     dict2=dict()
     for i in range(1, row_count_1 + 1):
          k=sheet_1.cell(row=i,column=column).value
          if k=="NA":
               continue
          else:
               data=int(str(k))
               if(data<= rank):
                    dict1[sheet_1['B'+str(i)].value]=data
     for i in range(1, row_count_2 + 1):
          v=sheet_2.cell(row=i,column=column).value
          if v=="NA":
               continue
          else:
               data=int(str(v)) 
               if(data<=rank):
                    dict2[sheet_2['B'+str(i)].value]=data
     
     keys = list(dict1.keys())
     values = list(dict1.values())
     sorted_value_index = np.argsort(values)
     sorted_dict_1 = {keys[i]: values[i] for i in sorted_value_index}

     keys1 = list(dict2.keys())
     values1 = list(dict2.values())
     sorted_value_index = np.argsort(values1)
     sorted_dict_2 = {keys1[i]: values1[i] for i in sorted_value_index}

     return [sorted_dict_1,sorted_dict_2]
